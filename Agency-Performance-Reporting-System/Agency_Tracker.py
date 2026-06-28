# %%
import sys
from pathlib import Path
from datetime import date
import pandas as pd
from openpyxl.reader.excel import load_workbook
#File Input & Output
input_file = Path(sys.argv[1]) if len(sys.argv)>1 else Path('final_raw.csv')
output_file= Path(sys.argv[2]) if len(sys.argv)>2 else Path(f'agency_performance_report_{date.today():%Y-%m-%d}.xlsx')


# %%
#Data Cleaning
import re
from datetime import datetime
from datetime import date, timedelta
import dateparser

def clean_data(df):
    def clean_name(name):
        text=str(name).lower()
        Titles = ['mr. ','mr ', 'ms. ', 'ms ', 'mrs. ','mrs ','mrs. ', 'miss ','dr. ', 'dr ','prof ', 'prof. ', 'phd ','phd']
        for title in Titles:
            text = text.replace(title, '')

        text = re.sub(r'\s+', ' ', text)
        return text.strip().title()

    df['Clean Name'] = df['Agent Name'].apply(clean_name)


    def clean_office(office):
        return office.strip().title()

    df['Clean Office'] = df['Office'].apply(clean_office)


    def clean_price(value):
        price =  str(value).replace('£','').replace(',','')
        return float(price)

    df['Clean Price'] =  df['Avg Sale Price'].apply(clean_price)



    def clean_date(value):
        value = re.sub(r'(\d+)(st|nd|rd|th)',r'\1', value)
        if '/' in value:
            try:
                return datetime.strptime(value,'%d/%m/%Y').date()
            except ValueError:
                return datetime.strptime(value,'%Y/%m/%d').date()
        elif '-' in value:
            try:
                return datetime.strptime(value,'%Y-%m-%d').date()
            except ValueError:
                return datetime.strptime(value,'%d-%m-%Y').date()

        else:
            parsed = dateparser.parse(value,settings={'DATE_ORDER': 'DMY'})
            return parsed.date() if parsed else None

    df['Clean Date'] =  df['Join Date'].apply(clean_date)

    df['Clean Specialisation'] =  df['Specialisation'].apply(lambda x: x.strip())
    df['Clean Contract'] = df['Contract'].apply(lambda x: x.strip())
    return df



# %%
#Calculations
def add_calculations(df):

    df['Customer Rating'] = df['Customer Rating'].apply(lambda x: float(x))

    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    df['Annual Revenue'] = df[months].sum(axis=1)


    df['Monthly Average'] = df[months].mean(axis=1).round(2)


    df['Best Month'] = df[months].idxmax(axis=1)
    df['Worst Month'] = df[months].idxmin(axis=1)

    df['Q1'] =df[['Jan','Feb','Mar']].sum(axis=1)
    df['Q2'] = df[['Apr','May','Jun']].sum(axis=1)
    df['Q3'] = df[['Jul','Aug','Sep']].sum(axis=1)
    df['Q4'] = df[['Oct','Nov','Dec']].sum(axis=1)


    df['Best Quarter'] =  df[['Q1','Q2','Q3','Q4']].idxmax(axis=1)
    df['Revenue vs Target'] = df['Annual Revenue'] - df['Annual Target']

    df['%Target'] =(df['Annual Revenue']/df['Annual Target']*100).round(1)

    def set_target(row):
        if row['%Target']>=110:
            return 'Exceeded'
        elif row['%Target']>=100:
            return 'Met'
        elif row['%Target']>=85:
            return 'Nearly Met'
        else:
            return 'Missed'

    df['Target Status'] = df.apply(set_target, axis=1)

    df['Conversion Rate'] = (df['Properties Sold']/df['Properties Listed']*100).round(1)

    df['Total Sale Value'] = df['Properties Sold'] * df['Clean Price']
    df['Total Sale Value'] = df['Total Sale Value'].round(0)
    df['Years Of Service'] = df['Clean Date'].apply(lambda x: (date.today() - x).days // 365 if x else None)

    df['Est Annual Hours'] = df['Clean Contract'].apply(lambda x: 1820 if x=='Full-Time' else 910)
    df['Est Annual Salary'] = df['Hourly Rate'] * df['Est Annual Hours']

    def calc_commission(row):
        specialisation= row['Specialisation']
        contract= row['Target Status']
        rev= row['Annual Revenue']

        if specialisation=='Luxury':
            if contract == 'Exceeded':
                return 0.15 * rev
            elif contract == 'Met':
                return 0.10 *rev
            else:
                return 0.05 *rev
        elif specialisation=='Commercial':
            if contract == 'Exceeded':
                return 0.10*rev
            elif contract == 'Met':
                return 0.07 * rev
            else:
                return 0.03*rev
        else:
            if contract == 'Exceeded':
                return 0.06 *rev
            elif contract == 'Met':
                return 0.04 * rev
            else:
                return 0.02*rev

    df['Commission'] =df.apply(calc_commission, axis=1)
    df['Commission'] = df['Commission'].round(0)

    def calc_performance(row):
        target = row['%Target']
        conversion = row['Conversion Rate']
        rating = row['Customer Rating']
        years = row['Years Of Service']
        points=0

        if target>=110:
            points = points +35
        elif target>=100:
            points = points +25
        elif target>=85:
            points = points +15
        if conversion>=70:
            points = points +25
        elif conversion>=50:
            points =points +15
        if rating>=4.5:
            points = points +20
        elif rating>=4.0:
            points = points +10
        if years>=5:
            points = points +10
        if points>100:
            points = 100
        return points

    df['Performance Score'] = df.apply(calc_performance, axis=1)

    def calc_agent(score):
        if score>=80:
            return 'Platinum'
        elif score>=60:
            return 'Gold'
        elif score>=40:
            return 'Silver'
        else:
            return 'Bronze'

    df['Agent Tier'] = df['Performance Score'].apply(calc_agent)

    def calc_alert (row):
        alerts= []
        if row['Conversion Rate']<50:
           alerts.append('LOW CONVERSION')
        if row['Target Status'] == 'Missed':
            alerts.append('MISSED TARGET')
        if row['Customer Rating']<3.5:
            alerts.append('LOW RATING')
        return ' | '.join(alerts) if alerts else ''

    df['Alert Flag'] = df.apply(calc_alert, axis=1)
    return df



# %%
#Summary Tables

def build_summaries(df):
    office_summary = df.groupby('Clean Office').agg(
        Agent_Count = ('AgentID','count'),
        Total_Revenue=('Annual Revenue','sum'),
        Target_Percentage = ('%Target','mean'),
        Total_Commission = ('Commission','sum'),
       Average_Performance_Score=('Performance Score','mean'),
        Platinum_Agents = ('Agent Tier', lambda x: (x=='Platinum').sum()),
        Bronze_Agents = ('Agent Tier', lambda x: (x=='Bronze').sum()),
    )
    office_summary = office_summary.round({
        'Target_Percentage':1,
        'Average_Performance_Score':1
    })


    manager_summary = df.groupby('Manager').agg(
        Agent_Count = ('AgentID','count'),
        Total_Revenue=('Annual Revenue','sum'),
        Total_Target=('Annual Target','sum'),
        Revenue_Vs_Target = ('Revenue vs Target','sum'),
        Total_Commission = ('Commission','sum'),
        Average_Performance_Score=('Performance Score','mean'),
        Total_Flags=('Alert Flag',lambda x:(x!="").sum()),


    )

    manager_summary = manager_summary.round({
            'Average_Performance_Score':1,
        })

    specialisation_summary = df.groupby('Specialisation').agg(
        Agent_Count = ('AgentID','count'),
        Total_Revenue=('Annual Revenue','sum'),
        Average_Sale_Price = ('Clean Price','mean'),
        Total_Commission = ('Commission','sum'),
        Average_Rating = ('Customer Rating','mean'),
       Exceeded_Target = ('Target Status',lambda x: (x == 'Exceeded').sum()),
    )

    specialisation_summary =specialisation_summary.round({
            'Average_Sale_Price':1,
            'Average_Rating':1
        })
    return office_summary, specialisation_summary,manager_summary
# %%
#Export & Style
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def header_style(df, Executive_Summary,office_summary, specialisation_summary, manager_summary):

    Flagged_Data=df[df['Alert Flag']!='']
    with pd.ExcelWriter(output_file) as writer:
        Executive_Summary.to_excel(writer, sheet_name='Executive Summary',index=False)
        df.to_excel(writer, sheet_name='Agent Data',index=False)
        office_summary.to_excel(writer, sheet_name='Office Summary',index=True)
        specialisation_summary.to_excel(writer, sheet_name='Specialisation Summary',index=True)
        manager_summary.to_excel(writer, sheet_name='Manager Summary',index=True)
        Flagged_Data.to_excel(writer, sheet_name='Flagged Data',index=False)

    wb = load_workbook(output_file)

    style_sheet(wb["Agent Data"])
    style_sheet(wb["Executive Summary"])
    style_sheet(wb["Office Summary"])
    style_sheet(wb["Specialisation Summary"])
    style_sheet(wb["Flagged Data"])
    style_sheet(wb["Manager Summary"])

    ws = wb["Executive Summary"]
    ws.column_dimensions["B"].width = 20

    wb.save(output_file)

def style_sheet(ws):

            for cell in ws[1]:
                cell.font = Font(bold=True,color='FFFFFF')
                cell.fill = PatternFill('solid', start_color='1F3864')
                cell.alignment = Alignment(horizontal='center')

            for col in ws.columns:
                max_len = max(len(str(cell.value or ''))for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4




def cond_format():
    from openpyxl import load_workbook
    from openpyxl.styles import  PatternFill
    def cond_fill(ws):
        headers ={}
        for cell in ws[1]:
            headers[cell.value]=cell.column_letter

        for cell in ws[headers['Agent Tier']][1:]:
            if cell.value == 'Platinum':
                cell.fill = PatternFill('solid', start_color='C6EFCE')
            elif cell.value== 'Bronze':
                cell.fill = PatternFill('solid', start_color='FFC7CE')
            elif cell.value not in (None,''):
                cell.fill = PatternFill('solid', start_color='FAC898')

    def cond_flagged_fill(ws):

      headers ={}
      for cell in ws[1]:
         headers[cell.value]=cell.column_letter

      for cell in ws[headers['Alert Flag']][1:]:
          if cell.value not in (None,''):
            cell.fill = PatternFill('solid', start_color='D0342C')

    wb = load_workbook(output_file)
    cond_fill(wb['Agent Data'])
    cond_fill(wb['Flagged Data'])
    cond_flagged_fill(wb['Flagged Data'])
    cond_flagged_fill(wb['Agent Data'])
    wb.save(output_file)


def int_format(ws, column_name):

    headers ={}
    for cell in ws[1]:
        headers[cell.value]=cell.column_letter

    if column_name in headers:
        for cell in ws[headers[column_name]][1:]:
            cell.number_format = '£#,##0'

def percent_format(ws, column_name):
    headers ={}
    for cell in ws[1]:
        headers[cell.value]=cell.column_letter

    if column_name in headers:
        for cell in ws[headers[column_name]][1:]:
            cell.number_format = '0.0"%"'

def specific_format(ws):
    headers ={}
    for cell in ws[1]:
        headers[cell.value]=cell.column_letter

    for row in range(2,ws.max_row+1):
        metric=ws[f'{headers['Metric']}{row}'].value
        value_cell= ws[f'{headers['Value']}{row}']

        if metric in('Total Revenue', 'Total Commission'):
            value_cell.number_format = '£#,##0.00'
        else:
            value_cell.number_format = '0'




# %%
#Terminal Summary

def print_summary(df):
    total_revenue = df['Annual Revenue'].sum()
    platinum_count =(df['Agent Tier']=='Platinum').sum()
    bronze_count = (df['Agent Tier']=='Bronze').sum()
    flagged_count = (df['Alert Flag']!='').sum()

    print(f'✅ Agency Report Generated - {len(df)} agents processed')
    print(f'💰 Total Revenue: £{total_revenue:,.0f}')
    print(f'🏆 Platinum Agents: {platinum_count}')
    print(f'🥉 Bronze Agents: {bronze_count}')
    print(f'🚩 Flagged Agents: {flagged_count}')
    print(f'📁 Saved to {output_file}')

# %%
#Executive Summary
def build_executive_summary(df):
    total_revenue = df['Annual Revenue'].sum()
    total_commission = df['Commission'].sum()
    platinum_count =(df['Agent Tier']=='Platinum').sum()
    flagged_count = (df['Alert Flag']!='').sum()
    avg_performance_score= (df['Performance Score'].mean()).round(1)
    Executive_Summary =  pd.DataFrame({
        'Metric':[
            'Total Revenue',
            'Total Commission',
            'Platinum Count',
            'Flagged Count',
            'Average Performance Score',
        ],
        'Value' : [
            total_revenue,
            total_commission,
            platinum_count,
            flagged_count,
            avg_performance_score
        ]
    })
    return Executive_Summary
# %%
if __name__ =='__main__':
    df = pd.read_csv(input_file)
    df.loc[df['AgentID'] == 'AG003', 'Manager'] = 'Sarah Mitchell'
    df.loc[df['AgentID'] == 'AG003', 'Specialisation'] = 'Residential'
    df = clean_data(df)
    df = add_calculations(df)
    Executive_Summary = build_executive_summary(df)
    office_summary, specialisation_summary, manager_summary = build_summaries(df)
    header_style(df,Executive_Summary, office_summary, specialisation_summary, manager_summary)
    cond_format()

    wb = load_workbook(output_file)

    specific_format(wb['Executive Summary'])

    int_format(wb['Office Summary'], 'Total_Revenue')
    int_format(wb['Office Summary'], 'Total_Commission')
    percent_format(wb['Office Summary'], 'Target_Percentage')

    int_format(wb['Manager Summary'], 'Total_Revenue')
    int_format(wb['Manager Summary'], 'Total_Target')
    int_format(wb['Manager Summary'], 'Revenue_Vs_Target')
    int_format(wb['Manager Summary'], 'Total_Commission')

    int_format(wb['Specialisation Summary'], 'Total_Revenue')
    int_format(wb['Specialisation Summary'], 'Average_Sale_Price')
    int_format(wb['Specialisation Summary'], 'Total_Commission')

    percent_format(wb['Agent Data'], '%Target')
    percent_format(wb['Agent Data'], 'Conversion Rate')
    percent_format(wb['Flagged Data'], '%Target')
    percent_format(wb['Flagged Data'], 'Conversion Rate')


    int_format(wb['Agent Data'], 'Annual Target')
    int_format(wb['Agent Data'], 'Clean Price')

    month_cols =  ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    quarter_cols = ['Q1','Q2','Q3','Q4']

    for col in month_cols + quarter_cols:
        int_format(wb['Agent Data'], col)
        int_format(wb['Flagged Data'], col)

    int_format(wb['Agent Data'], 'Annual Revenue')
    int_format(wb['Agent Data'], 'Monthly Average')
    int_format(wb['Agent Data'], 'Revenue vs Target')
    int_format(wb['Agent Data'], 'Commission')
    int_format(wb['Agent Data'], 'Total Sale Value')
    int_format(wb['Agent Data'], 'Est Annual Salary')

    int_format(wb['Flagged Data'], 'Annual Target')
    int_format(wb['Flagged Data'], 'Clean Price')
    int_format(wb['Flagged Data'], 'Annual Revenue')
    int_format(wb['Flagged Data'], 'Monthly Average')
    int_format(wb['Flagged Data'], 'Revenue vs Target')
    int_format(wb['Flagged Data'], 'Commission')
    int_format(wb['Flagged Data'], 'Total Sale Value')
    int_format(wb['Flagged Data'], 'Est Annual Salary')


    wb.save(output_file)
    print_summary(df)